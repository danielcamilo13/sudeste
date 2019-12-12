# -*- coding: latin -*-
from openpyxl import load_workbook,Workbook
import os,sys,re,csv

def reading(planilha):
    print(planilha)
    data = {}
    wb = load_workbook(filename=planilha)
    ws = wb.active
    for r in range(4,ws.max_row):
        for c in range(1,ws.max_column):
            if ws.cell(row=r,column=c).value is not None:
                nome = ws.cell(row=r,column=1).value
                cidade = ws.cell(row=r,column=2).value
                comprador = ws.cell(row=r,column=9).value
                # data.append[({'nome':nome,'cidade':cidade,'comprador':comprador})]
                print('dados {}'.format(data))
    return planilha

def spreadsheet_reader(spreadsheet):
    data = []
    wb = load_workbook(filename=spreadsheet)
    ws = wb.active
    last_row = ws.max_row
    last_column = ws.max_column
    for r in range(1,last_row):
        lines = []
        for c in range(1,last_column):
            v = ws.cell(row=r,column=c).value
            if v is None:
                v='-'
            lines.append(v)
        data.append(lines)
    return data
